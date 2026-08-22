import html
import openpyxl
import string
import cherrypy
from mako.template import Template
from spiderfoot import __version__

class WebUiHelpers:
    def cleanUserInput(self, inputList):
        if not isinstance(inputList, list):
            raise TypeError(f"inputList is {type(inputList)}; expected list()")
        ret = list()
        for item in inputList:
            if not item:
                ret.append("")
                continue
            c = html.escape(item, True)
            c = c.replace("&amp;", "&").replace("&quot;", "\"")
            ret.append(c)
        return ret

    def buildExcel(self, data, columnNames, sheetNameIndex=0):
        rowNums = dict()
        from sfwebui import BytesIO  # Local import to avoid circular import
        workbook = openpyxl.Workbook()
        defaultSheet = workbook.active
        columnNames = list(columnNames)  # Avoid mutating input
        columnNames.pop(sheetNameIndex)
        allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
        for row in data:
            row = list(row)
            sheetName = "".join([
                c for c in str(row.pop(sheetNameIndex)) if c.upper() in allowed_sheet_chars
            ])
            try:
                worksheet = workbook[sheetName]
            except KeyError:
                worksheet = workbook.create_sheet(sheetName)
                rowNums[sheetName] = 1
                for col_num, header in enumerate(columnNames, 1):
                    worksheet.cell(row=1, column=col_num, value=header)
                rowNums[sheetName] = 2
            for col_num, cell_value in enumerate(row, 1):
                worksheet.cell(row=rowNums[sheetName], column=col_num, value=str(cell_value))
            rowNums[sheetName] += 1
        if rowNums:
            workbook.remove(defaultSheet)
        workbook._sheets.sort(key=lambda ws: ws.title)
        with BytesIO() as f:
            workbook.save(f)
            f.seek(0)
            return f.read()

    def jsonify_error(self, status, message):
        cherrypy.response.headers['Content-Type'] = 'application/json'
        cherrypy.response.status = status
        return {
            'error': {
                'http_status': status,
                'message': message,
            }
        }

    def error(self, message):
        templ = Template(
            filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message=message, docroot=self.docroot, version=__version__)

    def searchBase(self, id=None, eventType=None, value=None):
        import time
        retdata = []
        if not id and not eventType and not value:
            return retdata
        if not value:
            value = ''
        regex = ""
        if value.startswith("/") and value.endswith("/"):
            regex = value[1:len(value) - 1]
            value = ""
        value = value.replace('*', '%')
        if value in [None, ""] and regex in [None, ""]:
            value = "%"
            regex = ""
        from spiderfoot import SpiderFootDb
        dbh = SpiderFootDb(self.config)
        criteria = {
            'scan_id': id or '',
            'type': eventType or '',
            'value': value or '',
            'regex': regex or '',
        }
        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata
        # dbh.search() returns (generated, data, source_data, module, type,
        # confidence, visibility, risk, hash, source_event_hash, event_descr,
        # event_type, scan_instance_id, false_positive, parent_false_positive)
        # -- searchResults() in scaninfo.tmpl reads output[0]=lastseen,
        # [1]=data, [2]=source data, [3]=source module, [8]=type. This used
        # to index past the end of dbh.search()'s row (up to row[14] against
        # a 9-column result), so every search with any matching data raised
        # IndexError, silently caught by search()'s bare except and returned
        # as an empty result. NOTE: scansearchresultexport()'s CSV/Excel
        # export expects a different 9-column order ("Date, Type, Value,
        # Source, Module, Risk, FP, Correlation, EventId") that this row
        # shape does NOT match -- a separate, pre-existing inconsistency
        # between the two consumers of this method, not introduced here.
        for row in data:
            lastseen = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            import html
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc,
                            row[3], row[9], row[5], row[6], row[7], row[4]])
        return retdata
