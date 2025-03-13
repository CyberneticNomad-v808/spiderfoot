"""
SpiderFoot FastAPI File Utilities

This module provides utilities for working with files in the FastAPI implementation.
"""

import csv
import json
import os
from io import StringIO, BytesIO
from typing import List, Dict, Any, Optional, Union, BinaryIO

import openpyxl
from fastapi.responses import Response, StreamingResponse
from fastapi import UploadFile


def create_csv_response(
    data: List[List[Any]], 
    filename: str, 
    headers: List[str] = None, 
    dialect: str = "excel"
) -> Response:
    """Create a CSV response.
    
    Args:
        data: Data rows
        filename: Output filename
        headers: CSV header row
        dialect: CSV dialect
    
    Returns:
        Response with CSV data
    """
    fileobj = StringIO()
    writer = csv.writer(fileobj, dialect=dialect)
    
    if headers:
        writer.writerow(headers)
    
    for row in data:
        writer.writerow(row)
    
    response_headers = {
        "Content-Disposition": f"attachment; filename={filename}",
        "Content-Type": "text/csv",
        "Pragma": "no-cache"
    }
    
    return Response(
        content=fileobj.getvalue().encode("utf-8"),
        headers=response_headers
    )


def create_excel_response(
    data: List[List[Any]], 
    filename: str, 
    headers: List[str] = None,
    sheet_name_index: int = 0
) -> Response:
    """Create an Excel response.
    
    Args:
        data: Data rows
        filename: Output filename
        headers: Excel header row
        sheet_name_index: Index of column to use for sheet names
    
    Returns:
        Response with Excel data
    """
    workbook = openpyxl.Workbook()
    
    if sheet_name_index >= 0:
        # Group data by sheet name
        sheet_data = {}
        sheet_headers = headers.copy() if headers else []
        
        if sheet_headers and sheet_name_index < len(sheet_headers):
            sheet_name_header = sheet_headers.pop(sheet_name_index)
        else:
            sheet_name_header = None
        
        for row in data:
            if sheet_name_index < len(row):
                sheet_name = str(row.pop(sheet_name_index))
                sheet_name = "".join(c for c in sheet_name if c.isalnum() or c == '_')
                
                if not sheet_name:
                    sheet_name = "Sheet1"
                
                if sheet_name not in sheet_data:
                    sheet_data[sheet_name] = []
                
                sheet_data[sheet_name].append(row)
        
        # Remove the default sheet
        if workbook.active and len(sheet_data) > 0:
            workbook.remove(workbook.active)
        
        # Create sheets and add data
        for sheet_name, rows in sheet_data.items():
            sheet = workbook.create_sheet(sheet_name)
            
            # Add headers
            if sheet_headers:
                for col_num, header in enumerate(sheet_headers, 1):
                    sheet.cell(row=1, column=col_num, value=header)
                
                # Add data rows
                for row_num, row_data in enumerate(rows, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        sheet.cell(row=row_num, column=col_num, value=cell_value)
            else:
                # Add data rows without headers
                for row_num, row_data in enumerate(rows, 1):
                    for col_num, cell_value in enumerate(row_data, 1):
                        sheet.cell(row=row_num, column=col_num, value=cell_value)
    else:
        # Use a single sheet
        sheet = workbook.active
        
        # Add headers
        if headers:
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)
            
            # Add data rows
            for row_num, row_data in enumerate(data, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_num, column=col_num, value=cell_value)
        else:
            # Add data rows without headers
            for row_num, row_data in enumerate(data, 1):
                for col_num, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_num, column=col_num, value=cell_value)
    
    # Save to bytes
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response_headers = {
        "Content-Disposition": f"attachment; filename={filename}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Pragma": "no-cache"
    }
    
    return Response(content=output.read(), headers=response_headers)


def process_uploaded_config_file(upload_file: UploadFile) -> Dict[str, str]:
    """Process an uploaded configuration file.
    
    Args:
        upload_file: Uploaded file
    
    Returns:
        Dictionary of configuration options
    """
    if not upload_file or not upload_file.file:
        return {}
    
    try:
        contents = upload_file.file.read()
        
        if isinstance(contents, bytes):
            contents = contents.decode("utf-8")
        
        config = {}
        for line in contents.split("\n"):
            if "=" not in line:
                continue
            
            opt_array = line.strip().split("=")
            if len(opt_array) == 1:
                opt_array.append("")
            
            config[opt_array[0]] = "=".join(opt_array[1:])
        
        return config
    except Exception as e:
        raise ValueError(f"Failed to process configuration file: {str(e)}")
